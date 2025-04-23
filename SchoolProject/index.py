import hashlib
import math

from flask import render_template, request, redirect, session, jsonify, flash, url_for, send_file
from flask_login import login_user, current_user, logout_user, login_required
import cloudinary.uploader
from sqlalchemy import text
from flask_wtf import FlaskForm
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

from SchoolProject import app, login, db
from SchoolProject.models import Admin, UserEnum, User, Grade, SchoolClass, Regulation, Subject, ScoreDetail
import dao

@app.context_processor
def utility_processor():
    def get_score(scores, score_type, attempt=None):
        if not scores:
            return ""
        for score in scores:
            if score_type == 'FINAL' and score.score_type == 'FINAL':
                return score.score_value
            elif score_type == '15P' and score.score_type == f'15P{attempt}':
                    return score.score_value
            elif score_type == '1T' and score.score_type == f'1T{attempt}':
                    return score.score_value
        return ""
    return dict(get_score=get_score)

def calculate_age(birthday_str):
    # Convert the string to a datetime object
    birth_date = datetime.strptime(birthday_str, "%Y-%m-%d").date()
    today = datetime.today().date()

    # Calculate age
    age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
    return age

@app.route('/')
def index():
    # Get search query
    q = request.args.get('q', '')
    tab = request.args.get('tab', 'grades')  # Default tab is grades
    
    # Load data based on active tab
    grades = []
    students = []
    if tab == 'grades':
        grades = dao.load_grade()
    elif tab == 'search':
        if q:
            students = dao.load_student(q=q)
            # Add class name for each student
            for student in students:
                if student.class_id:
                    class_info = dao.get_class_by_id(student.class_id)
                    if class_info:
                        student.class_name = class_info.name
                    else:
                        student.class_name = "Chưa phân lớp"
                else:
                    student.class_name = "Chưa phân lớp"
            
    return render_template("index.html", 
                         grades=grades,
                         students=students,
                         search_query=q,
                         active_tab=tab)

@app.route('/grade/<int:grade_id>')
def class1(grade_id):
    classes = dao.load_class(grade_id=grade_id)
    # Add student count for each class
    for c in classes:
        c.total_students = dao.count_class(class_id=c.id)
    return render_template("class.html", classes=classes, grade_id=grade_id)


@app.route('/grade/<int:grade_id>/class/<int:id>')
def inclass(grade_id, id):
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q")
    
    # Get total count first
    ccount = dao.count_class(class_id=id)
    
    # Calculate total pages
    page_size = app.config["PAGE_SIZE"]
    total_pages = math.ceil(ccount / page_size)
    
    # Get paginated students
    students = dao.load_student(q=q, class_id=id, page=page)
    
    # Get class info
    clazz = dao.get_class_by_id(id)
    
    return render_template("inclass.html", 
                           students=students,
                       clazz=clazz,
                           ccount=ccount,
                       current_page=page,
                       total_pages=total_pages,
                           grade_id=grade_id)

@app.route('/student/<int:id>')
def details(id):
    stu = dao.get_student_by_id(id)
    if stu and stu.class_id:
        class_info = dao.get_class_by_id(stu.class_id)
        if class_info:
            stu.class_name = class_info.name
        else:
            stu.class_name = "Chưa phân lớp"
    else:
        stu.class_name = "Chưa phân lớp"
    return render_template('student_details.html', stu=stu)

@login.user_loader
def get_user(user_id):
    return dao.get_user_by_id(user_id=user_id)

@app.route('/get-classes')
def get_classes():
    grade_id = request.args.get('grade_id')
    classes = dao.load_class(grade_id=grade_id)
    return jsonify([{'id': c.id, 'name': c.name} for c in classes])

@app.route("/registerstudent", methods=['GET','POST'])
def registerstudent():
    err_msg = None
    success_msg = None
    form = FlaskForm()  # Create form instance for CSRF token
    
    # Get or create regulations at the start
    regulations = Regulation.query.first()
    if not regulations:
        regulations = Regulation()
        db.session.add(regulations)
        db.session.commit()

    if request.method == "POST":
        if form.validate_on_submit():  # Check CSRF token
            try:
                name = request.form.get('name')
                avatar = request.files.get('avatar')
                path = None
                if avatar:
                    res = cloudinary.uploader.upload(avatar)
                    path = res['secure_url']
                birthday_str = request.form.get('birthday')
                
                # Validate required fields
                if not all([name, birthday_str]):
                    raise ValueError("Vui lòng điền đầy đủ họ tên và ngày sinh")
                
                # Calculate age and validate
                age = calculate_age(birthday_str)
                if age < regulations.min_age:
                    raise ValueError(f"Học sinh phải đủ {regulations.min_age} tuổi")
                if age > regulations.max_age:
                    raise ValueError(f"Học sinh không được quá {regulations.max_age} tuổi")
                
                gender = request.form.get('gender')
                address = request.form.get('address')
                phone = request.form.get('phone')
                email = request.form.get('email')
                if email and '@' not in email:
                    raise ValueError("Email không hợp lệ. Vui lòng nhập địa chỉ có dấu @.")
                class_id = request.form.get('class_id')
                if not class_id:
                    raise ValueError("Vui lòng chọn lớp cho học sinh.")
                    
                # Check class size limit
                current_count = dao.count_class(class_id)
                if current_count >= regulations.max_students_per_class:
                    raise ValueError(f"Lớp đã đủ {regulations.max_students_per_class} học sinh. Không thể thêm nữa.")
                
                success, message = dao.add_user(name, path, birthday_str, gender, address, phone, email, class_id)
                if success:
                    success_msg = message
                else:
                    err_msg = message
                    
            except Exception as e:
                err_msg = f"Lỗi: {str(e)}"
        else:
            err_msg = "CSRF token không hợp lệ hoặc đã hết hạn"
    
    grades = dao.load_grade()
    return render_template('register_student.html', 
                         err_msg=err_msg, 
                         success_msg=success_msg, 
                         grades=grades,
                         regulations=regulations,
                         form=form,
                         now=datetime.now())

@app.route("/registerteacher", methods=['GET', 'POST'])
@login_required
def register_teacher():
    err_msg = None
    success_msg = None

    if request.method == "POST":
        try:
            name = request.form.get('name')
            username = request.form.get('username')
            password = request.form.get('password')
            role = 2 if request.form.get('role') == '2' else 1  # 2 for TEACHER, 1 for ADMIN
            class_id = request.form.get('class_id') if role == 2 else None
            
            # Validate input
            if not all([name, username, password]):
                raise Exception("Vui lòng điền đầy đủ thông tin!")
                
            # Check username
            if Admin.query.filter_by(username=username).first():
                raise Exception("Tên đăng nhập đã tồn tại!")

            # Check class for teacher only
            if role == 2:  # TEACHER
                if not class_id:
                    raise Exception("Vui lòng chọn lớp cho giáo viên!")
                # Only check for duplicate class_id for teachers
                if Admin.query.filter_by(class_id=class_id, role=UserEnum.TEACHER.value).first():
                    raise Exception("Lớp này đã được phân cho một giáo viên khác!")
            
            # Create new admin/teacher
            success, message = dao.add_teacher(
                name=name, 
                username=username, 
                password=password,
                role=role,
                class_id=class_id
            )
            
            if success:
                success_msg = message
            else:
                err_msg = message

        except Exception as e:
            err_msg = str(e)
            db.session.rollback()

    classes = dao.load_class()
    # Get classes that already have teachers
    used_class_ids = [c.class_id for c in Admin.query.filter(
        Admin.class_id.isnot(None),
        Admin.role == UserEnum.TEACHER.value
    ).all()]
    
    return render_template("register_admin.html", 
                         classes=classes, 
                         used_class_ids=used_class_ids,
                         err_msg=err_msg, 
                         success_msg=success_msg)

@app.route('/login', methods=['get', 'post'])
def login_my_user():
    if current_user.is_authenticated:
        return redirect('/')

    err_msg = None
    next = None
    form = FlaskForm()  # Create form instance for CSRF token
    
    if request.method.__eq__('POST'):
        username = request.form.get('username')
        password = request.form.get('password')

        hashed_password = hashlib.md5(password.strip().encode('utf-8')).hexdigest()

        admin = Admin.query.filter_by(username=username.strip(), password=hashed_password).first()

        if admin:
            login_user(admin)
            next = request.args.get('next')
            return redirect(next if next else '/')
        else:
            err_msg = "Tài khoản hoặc mật khẩu không khớp!"

    return render_template('login.html', err_msg=err_msg, form=form)

@app.route('/teacher/class-list')
@login_required
def teacher_class_list():
    clazz = dao.get_class_by_id(current_user.class_id)
    students = dao.load_student(class_id=current_user.class_id)
    return render_template('teacher_class.html', clazz=clazz, students=students)

@app.route("/logout")
def logout_my_user():
    logout_user()
    return redirect('/login')


@app.route('/teacher/input-score/<int:student_id>', methods=['GET', 'POST'])
@login_required
def input_score_single(student_id):
    # 1. Kiểm tra quyền và lấy thông tin học sinh
    student = dao.get_student_by_id(student_id)
    if not student or student.class_id != current_user.class_id:
        return redirect('/teacher/class-list')

    # Lấy danh sách môn học đang active
    active_subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    if not active_subjects:
        flash('Chưa có môn học nào được kích hoạt', 'error')
        return redirect('/teacher/class-list')

    # Khởi tạo form_data với giá trị mặc định
    form_data = {
        'subject': request.args.get('subject') or active_subjects[0].name,
        'semester': request.args.get('semester', 1),
        'year': request.args.get('year', '2024-2025'),
    }

    # 2. Xử lý GET request - Hiển thị form với dữ liệu hiện có
    if request.method == 'GET':
        existing_scores = dao.get_scores_by_subject(
            student_id=student_id,
            subject=form_data['subject'],
            semester=form_data['semester'],
            year=form_data['year']
        )

        return render_template('add_score_single.html',
                               student=student,
                               existing_scores=existing_scores,
                            subjects=[s.name for s in active_subjects],
                            form_data=form_data)

    # 3. Xử lý POST request - Lưu điểm mới
    try:
        # Kiểm tra môn học có hợp lệ
        subject = request.form.get('subject')
        if not Subject.query.filter_by(name=subject, active=True).first():
            raise Exception('Môn học không hợp lệ hoặc đã bị vô hiệu hóa')

        # Cập nhật form_data với dữ liệu từ form
        form_data.update({
            'subject': subject,
            'semester': request.form.get('semester'),
            'year': request.form.get('year'),
            **{f'd15_{i}': request.form.get(f'd15_{i}') for i in range(1, 6)},
            **{f'd1t_{i}': request.form.get(f'd1t_{i}') for i in range(1, 4)},
            'final': request.form.get('final')
        })

        # Gọi DAO để lưu điểm
        success, message = dao.add_score(
            student_id=student_id,
            form_data=form_data
        )

        # 4. Sau khi lưu, hiển thị lại form với dữ liệu mới
        existing_scores = dao.get_scores_by_subject(
            student_id=student_id,
            subject=form_data['subject'],
            semester=form_data['semester'],
            year=form_data['year']
        )

        if success:
            return render_template('add_score_single.html',
                                   student=student,
                                   existing_scores=existing_scores,
                                subjects=[s.name for s in active_subjects],
                                success_msg=message,
                                form_data=form_data)
        else:
            return render_template('add_score_single.html',
                                   student=student,
                                   existing_scores=existing_scores,
                                subjects=[s.name for s in active_subjects],
                                error_msg=message,
                                form_data=form_data)

    except Exception as e:
        return render_template('add_score_single.html',
                               student=student,
                            subjects=[s.name for s in active_subjects],
                            error_msg=f"Lỗi hệ thống: {str(e)}",
                            form_data=form_data)

@app.route('/get-scores')
@login_required
def get_scores():
    student_id = request.args.get('student_id', type=int)
    subject = request.args.get('subject')
    semester = request.args.get('semester', type=int)
    year = request.args.get('year')

    # Kiểm tra quyền truy cập
    student = dao.get_student_by_id(student_id)
    if not student or student.class_id != current_user.class_id:
        return jsonify([])

    # Lấy điểm từ database
    scores = dao.get_scores_by_subject(
        student_id=student_id,
        subject=subject,
        semester=semester,
        year=year
    )

    # Chuyển đổi điểm sang JSON
    score_list = []
    for score in scores:
        score_list.append({
            'type': score.score_type,
            'value': score.score_value
        })

    return jsonify(score_list)

@app.route('/class/<int:class_id>/scores')
@login_required
def class_scores(class_id):
    # Get parameters
    subject = request.args.get('subject')
    semester = request.args.get('semester', 1, type=int)
    year = request.args.get('year', '2024-2025')
    
    # Get class info
    clazz = dao.get_class_by_id(class_id)
    if not clazz:
        return redirect('/')
        
    # Get active subjects
    subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    if not subject and subjects:
        subject = subjects[0].name
    
    # Get average scores
    averages = dao.get_class_average_scores(class_id, subject, semester, year)
    
    # Calculate class average
    class_average = sum(s['average'] for s in averages) / len(averages) if averages else 0
    
    return render_template('class_scores.html',
                         class_name=clazz.name,
                         subject=subject,
                         subjects=subjects,
                         semester=semester,
                         year=year,
                         averages=averages,
                         class_average=class_average)

@app.route('/class/<int:class_id>/all-scores')
@login_required
def class_all_scores(class_id):
    # Get parameters
    semester = request.args.get('semester', 1, type=int)
    year = request.args.get('year', '2024-2025')
    
    # Get class info
    clazz = dao.get_class_by_id(class_id)
    if not clazz:
        return redirect('/')
    
    # Get all students in the class
    students = dao.load_student(class_id=class_id)
    
    # Get active subjects
    subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    subject_names = [s.name for s in subjects]
    
    # Calculate averages for each student and subject
    class_total_average = 0
    valid_student_count = 0
    
    for student in students:
        student.averages = {}
        total_score = 0
        valid_subjects = 0
        
        for subject in subject_names:
            avg = dao.calculate_average_score(student.id, subject, semester, year)
            if avg is not None:
                student.averages[subject] = {'average': avg}
                total_score += avg
                valid_subjects += 1
        
        # Calculate overall average for student
        if valid_subjects > 0:
            student.overall_average = round(total_score / valid_subjects, 1)
            class_total_average += student.overall_average
            valid_student_count += 1
        else:
            student.overall_average = None
    
    # Calculate class averages for each subject and overall
    class_averages = {}
    for subject in subject_names:
        subject_scores = [s.averages.get(subject, {}).get('average') 
                         for s in students 
                         if s.averages.get(subject, {}).get('average') is not None]
        if subject_scores:
            class_averages[subject] = sum(subject_scores) / len(subject_scores)
    
    # Calculate class overall average
    class_overall_average = round(class_total_average / valid_student_count, 1) if valid_student_count > 0 else None
    
    return render_template('class_all_scores.html',
                         class_name=clazz.name,
                         semester=semester,
                         year=year,
                         students=students,
                         subjects=subject_names,
                         class_averages=class_averages,
                         class_overall_average=class_overall_average)

@app.route('/admin/subject-stats')
@login_required
def admin_subject_stats():
    # Set up academic years
    current_year = 2024
    prev_year = current_year - 1
    current_year_str = f"{current_year}-{current_year + 1}"
    prev_year_str = f"{prev_year}-{current_year}"
    
    # Get active subjects
    subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    if not subjects:
        flash('Chưa có môn học nào được kích hoạt', 'warning')
        return redirect(url_for('manage_regulations'))
    
    # Get selected subject (default to first active subject)
    subject = request.args.get('subject')
    if not subject and subjects:
        subject = subjects[0].name
    
    semester = int(request.args.get('semester', '1'))
    year = request.args.get('year', current_year_str)
    
    # Get all classes
    classes = dao.get_all_classes()
    class_stats = []
    
    for class_info in classes:
        # Get all students in the class
        students = dao.load_student(class_id=class_info['id'])
        total_students = len(students)
        passing_students = 0
        
        # Calculate passing students (average >= 5)
        for student in students:
            avg = dao.calculate_average_score(student.id, subject, semester, year)
            if avg and avg >= 5:
                passing_students += 1
        
        # Calculate passing rate
        passing_rate = (passing_students / total_students * 100) if total_students > 0 else 0
        
        class_stats.append({
            'class_name': class_info['name'],
            'total_students': total_students,
            'passing_students': passing_students,
            'passing_rate': passing_rate
        })
    
    return render_template('admin_subject_stats.html',
                         subjects=subjects,
                         subject=subject,
                         semester=semester,
                         year=year,
                         current_year_str=current_year_str,
                         prev_year_str=prev_year_str,
                         class_stats=class_stats)

@app.route('/admin/add-class', methods=['GET', 'POST'])
@login_required
def add_class():
    if not current_user.is_admin:
        return redirect('/')
        
    err_msg = None
    success_msg = None
    form = FlaskForm()
    
    if request.method == 'POST' and form.validate_on_submit():
        try:
            grade_id = request.form.get('grade_id')
            if not grade_id:
                raise Exception("Vui lòng chọn khối!")
                
            # Get grade info
            grade = Grade.query.get(grade_id)
            if not grade:
                raise Exception("Khối không tồn tại!")
                
            # Get existing classes for this grade
            existing_classes = SchoolClass.query.filter_by(grade_id=grade_id).all()
            
            # Add 'A' to grade name if it doesn't end with a letter
            base_name = grade.name
            if not base_name[-1].isalpha():
                base_name = f"{base_name}A"
            
            # Extract the base name (e.g., "11A") and find the highest number
            if not existing_classes:
                new_class_name = f"{base_name}1"  # e.g., "11A1" for first class
            else:
                # Find highest number
                max_num = 0
                for c in existing_classes:
                    try:
                        # Remove the grade number and 'A' to get the class number
                        class_num_str = c.name[len(base_name):]
                        if class_num_str:  # Make sure there's a number
                            num = int(class_num_str)
                            max_num = max(max_num, num)
                    except ValueError:
                        continue
                new_class_name = f"{base_name}{max_num + 1}"
            
            # Add class
            dao.add_class(name=new_class_name, grade_id=grade_id)
            success_msg = f"Thêm lớp {new_class_name} thành công!"
            
        except Exception as e:
            err_msg = str(e)
    elif request.method == 'POST':
        err_msg = "CSRF token không hợp lệ"
    
    # Get grades with their classes
    grades = Grade.query.all()
    for grade in grades:
        grade.classes = SchoolClass.query.filter_by(grade_id=grade.id).all()
        for class_ in grade.classes:
            class_.students = User.query.filter_by(class_id=class_.id).all()
    
    return render_template('add_class.html', 
                         grades=grades,
                         err_msg=err_msg,
                         success_msg=success_msg,
                         form=form)

@app.route('/admin/delete-class/<int:class_id>', methods=['POST'])
@login_required
def delete_class(class_id):
    try:
        # Kiểm tra lớp tồn tại và lấy thông tin cần thiết
        result = db.session.execute(
            text("SELECT id, grade_id, name FROM classes WHERE id = :class_id"),
            {"class_id": class_id}
        ).fetchone()
        
        if not result:
            return jsonify({'success': False, 'message': 'Lớp không tồn tại!'}), 404
            
        grade_id = result[1]
        class_name = result[2]
        
        # Kiểm tra có phải lớp cuối cùng của khối
        class_count = db.session.execute(
            text("SELECT COUNT(*) FROM classes WHERE grade_id = :grade_id"),
            {"grade_id": grade_id}
        ).scalar()
        
        if class_count <= 1:
            return jsonify({'success': False, 'message': 'Không thể xóa lớp cuối cùng của khối!'}), 400
            
        # Lấy base name (e.g., "10A")
        base_name = class_name
        for i in range(len(base_name)-1, -1, -1):
            if not base_name[i].isdigit():
                base_name = base_name[:i+1]
                break
                
        deleted_number = int(class_name[len(base_name):])
        
        try:
            # Xóa học sinh trong lớp
            db.session.execute(
                text("DELETE FROM students WHERE class_id = :class_id"),
                {"class_id": class_id}
            )
            
            # Xóa lớp
            db.session.execute(
                text("DELETE FROM classes WHERE id = :class_id"),
                {"class_id": class_id}
            )
            
            # Lấy danh sách lớp cần đổi tên
            classes = db.session.execute(
                text("""
                SELECT id, name 
                FROM classes 
                WHERE grade_id = :grade_id 
                AND name LIKE :name_pattern
                ORDER BY name
                """),
                {
                    "grade_id": grade_id,
                    "name_pattern": f"{base_name}%"
                }
            ).fetchall()
            
            # Cập nhật tên lớp
            for class_row in classes:
                try:
                    current_number = int(class_row[1][len(base_name):])
                    if current_number > deleted_number:
                        new_name = f"{base_name}{current_number - 1}"
                        db.session.execute(
                            text("UPDATE classes SET name = :new_name WHERE id = :class_id"),
                            {"new_name": new_name, "class_id": class_row[0]}
                        )
                except ValueError:
                    continue
            
            # Commit các thay đổi
            db.session.commit()
            return jsonify({'success': True, 'message': 'Xóa lớp và cập nhật tên lớp thành công!'})
            
        except Exception as e:
            db.session.rollback()
            raise e
            
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

class DeleteForm(FlaskForm):
    pass

@app.route('/admin/class-management')
@login_required
def admin_class_management():
    if not current_user.is_admin:
        return redirect('/')
        
    # Get all grades with their classes and students
    grades = Grade.query.all()
    for grade in grades:
        grade.classes = SchoolClass.query.filter_by(grade_id=grade.id).all()
        for class_ in grade.classes:
            class_.students = User.query.filter_by(class_id=class_.id).all()
    
    form = DeleteForm()
    return render_template('admin_class_management.html', grades=grades, form=form)

@app.route('/admin/remove-student/<int:student_id>', methods=['POST'])
@login_required
def remove_student(student_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Không có quyền truy cập!'}), 403
        
    try:
        student = User.query.get(student_id)
        if not student:
            return jsonify({'success': False, 'message': 'Không tìm thấy học sinh!'}), 404
            
        # Lưu thông tin lớp cũ
        old_class = student.class_id
        
        # Xóa học sinh khỏi lớp (set class_id = None)
        student.class_id = None
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Đã xóa học sinh khỏi lớp thành công!'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/regulations', methods=['GET', 'POST'])
@login_required
def manage_regulations():
    if not current_user.is_admin:
        return redirect('/')
        
    err_msg = None
    success_msg = None
    
    # Get or create regulations
    regulations = Regulation.query.first()
    if not regulations:
        regulations = Regulation()
        db.session.add(regulations)
        db.session.commit()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'update_regulations':
            try:
                min_age = int(request.form.get('min_age'))
                max_age = int(request.form.get('max_age'))
                max_students = int(request.form.get('max_students_per_class'))
                
                if min_age >= max_age:
                    raise ValueError("Tuổi tối thiểu phải nhỏ hơn tuổi tối đa")
                    
                if max_students < 1:
                    raise ValueError("Sĩ số tối đa phải lớn hơn 0")
                    
                regulations.min_age = min_age
                regulations.max_age = max_age
                regulations.max_students_per_class = max_students
                
                db.session.commit()
                success_msg = "Cập nhật quy định thành công!"
            except ValueError as e:
                err_msg = str(e)
            except Exception as e:
                err_msg = f"Lỗi: {str(e)}"
                db.session.rollback()
            
        elif action == 'add_subject':
            subject_name = request.form.get('subject_name').strip()
            if not subject_name:
                raise ValueError("Tên môn học không được để trống")
                
            # Check if subject already exists
            if Subject.query.filter_by(name=subject_name).first():
                raise ValueError("Môn học này đã tồn tại")
                
            subject = Subject(name=subject_name)
            db.session.add(subject)
            db.session.commit()
            success_msg = f"Đã thêm môn học {subject_name}"
            
        elif action == 'update_subject':
            subject_id = request.form.get('subject_id')
            subject_name = request.form.get('subject_name')
            
            if not subject_id or not subject_name:
                flash('Thiếu thông tin cập nhật môn học', 'error')
                return redirect(url_for('manage_regulations'))
                
            subject = Subject.query.get(subject_id)
            if not subject:
                flash('Không tìm thấy môn học', 'error')
                return redirect(url_for('manage_regulations'))
                
            # Kiểm tra xem tên môn học mới có bị trùng không
            existing_subject = Subject.query.filter(
                Subject.name == subject_name,
                Subject.id != subject_id
            ).first()
            
            if existing_subject:
                flash('Tên môn học đã tồn tại', 'error')
                return redirect(url_for('manage_regulations'))
                
            try:
                subject.name = subject_name
                db.session.commit()
                flash('Cập nhật tên môn học thành công', 'success')
            except:
                db.session.rollback()
                flash('Có lỗi xảy ra khi cập nhật môn học', 'error')
            
            return redirect(url_for('manage_regulations'))
            
        elif action == 'delete_subject':
            subject_id = request.form.get('subject_id')
            subject = Subject.query.get(subject_id)
            if subject:
                # Check if subject has any scores
                if ScoreDetail.query.filter_by(subject=subject.name).first():
                    raise ValueError("Không thể xóa môn học này vì đã có điểm số")
                db.session.delete(subject)
                db.session.commit()
                success_msg = f"Đã xóa môn học {subject.name}"
                
        elif action == 'activate_subject':
            subject_id = request.form.get('subject_id')
            subject = Subject.query.get(subject_id)
            if subject:
                subject.is_active = True
                db.session.commit()
                success_msg = f"Đã kích hoạt môn học {subject.name}"
                
        elif action == 'deactivate_subject':
            subject_id = request.form.get('subject_id')
            subject = Subject.query.get(subject_id)
            if subject:
                subject.is_active = False
                db.session.commit()
                success_msg = f"Đã vô hiệu hóa môn học {subject.name}"
                
    # Get all subjects
    subjects = Subject.query.order_by(Subject.name).all()
            
    form = FlaskForm()
    return render_template('regulations.html', 
                         regulations=regulations,
                         subjects=subjects,
                         form=form,
                         error_msg=err_msg,
                         success_msg=success_msg)

@app.route('/admin/remove-teacher/<int:class_id>', methods=['POST'])
@login_required
def remove_teacher(class_id):
    if not current_user.is_admin:
        return jsonify({'success': False, 'message': 'Không có quyền thực hiện'}), 403
        
    try:
        # Tìm giáo viên phụ trách lớp này
        teacher = Admin.query.filter_by(class_id=class_id, role=2).first()
        if not teacher:
            return jsonify({'success': False, 'message': 'Không tìm thấy giáo viên phụ trách lớp này'}), 404
            
        # Xóa hoàn toàn tài khoản giáo viên khỏi database
        db.session.delete(teacher)
        db.session.commit()
        
        return jsonify({'success': True, 'message': 'Đã xóa giáo viên thành công'})
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500

@app.route('/admin/class-teachers')
@login_required
def manage_class_teachers():
    if not current_user.is_admin:
        return redirect(url_for('index'))
        
    classes = SchoolClass.query.order_by(SchoolClass.name).all()
    return render_template('admin_class_teacher.html', classes=classes)

@app.route('/admin/edit-student/<int:student_id>', methods=['GET', 'POST'])
@login_required
def edit_student(student_id):
    if not current_user.is_admin:
        return redirect('/')
        
    student = dao.get_student_by_id(student_id)
    if not student:
        flash('Không tìm thấy học sinh', 'error')
        return redirect('/')
    
    # Get or create regulations at the start
    regulations = Regulation.query.first()
    if not regulations:
        regulations = Regulation()
        db.session.add(regulations)
        db.session.commit()
        
    err_msg = None
    success_msg = None
    form = FlaskForm()
    
    if request.method == 'POST':
        if form.validate_on_submit():
            try:
                name = request.form.get('name')
                avatar = request.files.get('avatar')
                if avatar:
                    res = cloudinary.uploader.upload(avatar)
                    student.avatar = res['secure_url']
                    
                birthday_str = request.form.get('birthday')
                age = calculate_age(birthday_str)
                
                # Validate age against regulations
                if age < regulations.min_age:
                    raise ValueError(f"Học sinh phải đủ {regulations.min_age} tuổi")
                if age > regulations.max_age:
                    raise ValueError(f"Học sinh không được quá {regulations.max_age} tuổi")
                    
                gender = request.form.get('gender')
                address = request.form.get('address')
                phone = request.form.get('phone')
                email = request.form.get('email')
                class_id = request.form.get('class_id')
                
                # Validate input
                if not all([name, birthday_str, gender, address, email]):
                    raise ValueError("Vui lòng điền đầy đủ thông tin bắt buộc!")
                    
                if '@' not in email:
                    raise ValueError("Email không hợp lệ!")
                    
                # Update student info
                student.name = name
                student.birthday = age
                student.gender = gender
                student.address = address
                student.phone = phone
                student.email = email
                
                # Update class if changed
                if class_id and int(class_id) != student.class_id:
                    # Check class size limit
                    current_count = dao.count_class(class_id)
                    if current_count >= regulations.max_students_per_class:
                        raise ValueError(f"Lớp đã đủ {regulations.max_students_per_class} học sinh!")
                    student.class_id = class_id
                    
                db.session.commit()
                success_msg = "Cập nhật thông tin học sinh thành công!"
                
            except Exception as e:
                err_msg = str(e)
                db.session.rollback()
        else:
            err_msg = "CSRF token không hợp lệ"
            
    # Format birthday for the form
    if student.birthday:
        if isinstance(student.birthday, str):
            try:
                student.birthday = datetime.strptime(student.birthday, '%Y-%m-%d').date()
            except ValueError:
                pass
            
    # Get all grades and classes for the form
    grades = dao.load_grade()
    return render_template('edit_student.html', regulations=regulations,
                         student=student,
                         grades=grades,
                         err_msg=err_msg,
                         success_msg=success_msg,
                         form=form)

@app.route('/class/<int:class_id>/export-scores')
@login_required
def export_class_scores(class_id):
    # Get parameters
    semester = request.args.get('semester', 1, type=int)
    year = request.args.get('year', '2024-2025')
    
    # Get class info
    clazz = dao.get_class_by_id(class_id)
    if not clazz:
        return redirect('/')
    
    # Get all students in the class
    students = dao.load_student(class_id=class_id)
    
    # Get active subjects
    subjects = Subject.query.filter_by(active=True).order_by(Subject.name).all()
    subject_names = [s.name for s in subjects]
    
    # Calculate averages for each student and subject
    class_total_average = 0
    valid_student_count = 0
    
    for student in students:
        student.averages = {}
        total_score = 0
        valid_subjects = 0
        
        for subject in subject_names:
            avg = dao.calculate_average_score(student.id, subject, semester, year)
            if avg is not None:
                student.averages[subject] = {'average': avg}
                total_score += avg
                valid_subjects += 1
        
        # Calculate overall average for student
        if valid_subjects > 0:
            student.overall_average = round(total_score / valid_subjects, 1)
            class_total_average += student.overall_average
            valid_student_count += 1
        else:
            student.overall_average = None
    
    # Calculate class averages for each subject
    class_averages = {}
    for subject in subject_names:
        subject_scores = [s.averages.get(subject, {}).get('average') 
                         for s in students 
                         if s.averages.get(subject, {}).get('average') is not None]
        if subject_scores:
            class_averages[subject] = sum(subject_scores) / len(subject_scores)
    
    # Calculate class overall average
    class_overall_average = round(class_total_average / valid_student_count, 1) if valid_student_count > 0 else None

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Điểm TB Lớp {clazz.name}"

    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    center_alignment = Alignment(horizontal='center')

    # Write headers
    ws['A1'] = "BẢNG ĐIỂM TRUNG BÌNH LỚP"
    ws.merge_cells('A1:' + get_column_letter(len(subject_names) + 3) + '1')
    ws['A1'].alignment = center_alignment
    ws['A1'].font = Font(bold=True, size=14)

    # Write class info
    ws['A2'] = f"Lớp: {clazz.name}"
    ws['A3'] = f"Học kỳ: {semester}"
    ws['B3'] = f"Năm học: {year}"

    # Column headers
    headers = ['STT', 'Họ và tên'] + subject_names + ['Điểm TB']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        ws.column_dimensions[get_column_letter(col)].width = 15

    # Write student data
    for row, student in enumerate(students, 5):
        ws.cell(row=row, column=1, value=row-4).alignment = center_alignment
        ws.cell(row=row, column=2, value=student.name)
        
        for col, subject in enumerate(subject_names, 3):
            avg = student.averages.get(subject, {}).get('average')
            if avg is not None:
                cell = ws.cell(row=row, column=col, value=round(avg, 1))
                cell.alignment = center_alignment
            else:
                ws.cell(row=row, column=col, value="-").alignment = center_alignment
        
        # Write overall average
        last_col = len(subject_names) + 3
        if student.overall_average is not None:
            cell = ws.cell(row=row, column=last_col, value=student.overall_average)
            cell.alignment = center_alignment
            cell.font = Font(bold=True)
        else:
            ws.cell(row=row, column=last_col, value="-").alignment = center_alignment

    # Write class averages
    last_row = len(students) + 5
    ws.cell(row=last_row, column=1, value="").alignment = center_alignment
    ws.cell(row=last_row, column=2, value="Trung bình lớp").font = header_font
    
    for col, subject in enumerate(subject_names, 3):
        avg = class_averages.get(subject)
        if avg is not None:
            cell = ws.cell(row=last_row, column=col, value=round(avg, 1))
            cell.alignment = center_alignment
            cell.font = header_font
        else:
            ws.cell(row=last_row, column=col, value="-").alignment = center_alignment
    
    # Write class overall average
    last_col = len(subject_names) + 3
    if class_overall_average is not None:
        cell = ws.cell(row=last_row, column=last_col, value=class_overall_average)
        cell.alignment = center_alignment
        cell.font = header_font
    else:
        ws.cell(row=last_row, column=last_col, value="-").alignment = center_alignment

    # Save file
    filename = f"diem_tb_lop_{clazz.name}_{semester}_{year}.xlsx"
    filepath = os.path.join(app.root_path, 'static', 'exports', filename)
    
    # Create exports directory if it doesn't exist
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    
    wb.save(filepath)
    
    return send_file(
        filepath,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == "__main__":
    with app.app_context():
        app.run(debug=True)
